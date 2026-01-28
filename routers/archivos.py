import io
import os
import json
import shutil
import sqlite3
import zipfile
import tempfile
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX
from fastapi import APIRouter, UploadFile, File, Form
from typing import List

router = APIRouter()

# ==========================================
# FUNCIONES AUXILIARES DE COLOR (REFORZADAS)
# ==========================================

def get_color_from_cell(cell):
    """
    Extrae el color hexadecimal real de una celda.
    Maneja RGB (ARGB), Colores Indexados y Temas de Excel.
    """
    try:
        if not cell.fill or not hasattr(cell.fill, 'start_color'):
            return None
        
        sc = cell.fill.start_color
        
        # 1. Caso: RGB directo (frecuente en colores personalizados como 00B050)
        if sc.type == 'rgb' and sc.rgb and isinstance(sc.rgb, str):
            color = str(sc.rgb).upper()
            return color[-6:] if len(color) >= 6 else color
        
        # 2. Caso: Color Indexado
        if sc.type == 'indexed' and sc.indexed is not None:
            try:
                idx_color = COLOR_INDEX[sc.indexed]
                return str(idx_color).upper()[-6:]
            except:
                pass
        
        # 3. Caso: Temas (Fallback)
        if sc.type == 'theme':
            return f"THEME_{sc.theme}"
            
        return None
    except:
        return None

def hex_to_rgb(hex_code):
    if not hex_code or "THEME" in str(hex_code):
        return None
    try:
        h = str(hex_code).strip().upper()
        if h.startswith('#'): h = h[1:]
        if len(h) != 6: return None
        return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))
    except:
        return None

def color_distance(c1, c2):
    if not c1 or not c2: return 999
    return sum((a - b) ** 2 for a, b in zip(c1, c2)) ** 0.5

def interpretar_stock_por_color(cell):
    """
    Traduce el color de la celda a un estado de stock.
    """
    raw_color = get_color_from_cell(cell)
    if not raw_color:
        return "NO DEFINIDO", "SIN_COLOR"

    rgb = hex_to_rgb(raw_color)
    
    if not rgb:
        if raw_color in ["THEME_4", "THEME_8"]: return "HAY STOCK", raw_color
        if raw_color in ["THEME_5", "THEME_9"]: return "PREGUNTAR", raw_color
        if raw_color in ["THEME_6", "THEME_7"]: return "NO HAY STOCK", raw_color
        return "DESCONOCIDO", raw_color

    targets = {
        "HAY STOCK": [(0, 176, 80), (0, 255, 0), (146, 208, 80), (0, 128, 0)],
        "PREGUNTAR": [(255, 255, 0), (255, 192, 0), (255, 230, 0), (255, 255, 153)],
        "NO HAY STOCK": [(255, 0, 0), (192, 0, 0), (255, 102, 102), (255, 199, 206)]
    }
    
    TOLERANCIA = 150
    for estado, colores_rgb in targets.items():
        for target_rgb in colores_rgb:
            if color_distance(rgb, target_rgb) < TOLERANCIA:
                return estado, raw_color

    return "DESCONOCIDO", raw_color

# ==========================================
# ENDPOINT: PROCESAR INVENTARIO (POR PASOS)
# ==========================================

@router.post("/procesar-inventario-completo/")
async def procesar_inventario_completo(file: UploadFile = File(...)):
    """
    PASO 1: Leer todas las hojas e identificar columnas clave.
    La Columna A (índice 0) se define como la de códigos por defecto.
    PASO 2: Limpiar valores y procesar colores en la columna STOCK.
    PASO 3: Consolidar todo.
    """
    try:
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents), data_only=True)
        
        inventario_por_hoja = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row < 2: continue

            # --- IDENTIFICACIÓN DE COLUMNAS ---
            headers = []
            idx_stock = -1
            idx_codigo = 0 # Forzamos que la Columna A sea siempre la de códigos
            
            for i, cell in enumerate(ws[1]):
                val = str(cell.value).strip().upper() if cell.value else f"COL_{i}"
                headers.append(val)
                
                if val == "STOCK":
                    idx_stock = i

            items_hoja = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if not any(c.value is not None for c in row):
                    continue

                item = {
                    "ORIGEN_HOJA": sheet_name,
                    "FILA_EXCEL": row_idx
                }

                for idx, cell in enumerate(row):
                    if idx >= len(headers): break
                    nombre_col = headers[idx]
                    valor = cell.value

                    # Limpiar espacios en la columna A (Códigos)
                    if idx == idx_codigo:
                        if valor is not None:
                            valor = str(valor).strip()
                    
                    item[nombre_col] = valor

                    # Procesar color en columna STOCK (si se encuentra)
                    if idx == idx_stock:
                        estado, raw_hex = interpretar_stock_por_color(cell)
                        item["STOCK_ESTADO"] = estado
                        item["STOCK_COLOR_RAW"] = raw_hex

                items_hoja.append(item)
            
            inventario_por_hoja[sheet_name] = items_hoja

        # --- UNIFICACIÓN ---
        lista_consolidada = []
        for items in inventario_por_hoja.values():
            lista_consolidada.extend(items)

        return {
            "archivo": file.filename,
            "resumen": {hoja: len(items) for hoja, items in inventario_por_hoja.items()},
            "total_items": len(lista_consolidada),
            "datos": lista_consolidada
        }

    except Exception as e:
        import traceback
        return {"error": str(e), "trace": traceback.format_exc()}

# Endpoints originales omitidos para brevedad pero mantenidos en el archivo real

@router.post("/leer-excel/")
async def leer_excel(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents), data_only=True)
        resultado = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row < 2: continue
            headers = [str(c.value) for c in ws[1]]
            hoja_datos = []
            for row in ws.iter_rows(min_row=2):
                fila = {}
                for idx, cell in enumerate(row):
                    if idx >= len(headers): break
                    fila[headers[idx]] = cell.value
                hoja_datos.append(fila)
            resultado[sheet_name] = hoja_datos
        return resultado
    except Exception as e:
        return {"error": str(e)}

@router.post("/procesar-zip-sqlite/")
async def procesar_zip_sqlite(file: UploadFile = File(...), codigos: str = Form(...), codigosProveedor: str = Form(...)):
    temp_dir = tempfile.mkdtemp()
    zip_path = os.path.join(temp_dir, "archivo.zip")
    try:
        lista_codigos = json.loads(codigos) if codigos else []
        lista_codigos_prov = json.loads(codigosProveedor) if codigosProveedor else []
        with open(zip_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        db_path = None
        with zipfile.ZipFile(zip_path, 'r') as zf:
            zf.extractall(temp_dir)
            for f in zf.namelist():
                if f.endswith(('.sqlite', '.db', '.sqlite3')):
                    db_path = os.path.join(temp_dir, f)
                    break
        if not db_path: return {"error": "No hay base de datos"}
        conn = sqlite3.connect(db_path)
        dfs = []
        if lista_codigos:
            p = ','.join('?' for _ in lista_codigos)
            dfs.append(pd.read_sql_query(f"SELECT * FROM Articulos WHERE Codigo IN ({p})", conn, params=lista_codigos))
        if lista_codigos_prov:
            p = ','.join('?' for _ in lista_codigos_prov)
            dfs.append(pd.read_sql_query(f"SELECT * FROM Articulos WHERE CodigoParticular IN ({p})", conn, params=lista_codigos_prov))
        resultado = pd.concat(dfs).drop_duplicates().to_dict(orient="records") if dfs else []
        conn.close()
        return {"mensaje": "Éxito", "total": len(resultado), "datos": resultado}
    except Exception as e:
        return {"error": str(e)}
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)