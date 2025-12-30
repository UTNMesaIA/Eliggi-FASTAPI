import os
import io
import json
import shutil
import sqlite3
import zipfile
import tempfile
from typing import List, Optional

# Imports de terceros
from fastapi import FastAPI, UploadFile, File, BackgroundTasks, HTTPException, Form
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from openpyxl import load_workbook
import pandas as pd
from sqlalchemy import create_engine, Column, Integer, String, MetaData, Table
from sqlalchemy.orm import sessionmaker

# ==========================================
# 1. CONFIGURACIÓN INICIAL Y DB (POSTGRES)
# ==========================================

app = FastAPI()

# Configuración CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configuración de Postgres (Para /upload-sheet)
DB_PASSWORD = os.getenv("PGPASSWORD")
DB_HOST = "gondola.proxy.rlwy.net"
DB_PORT = "43938"
DB_USER = "postgres"
DB_NAME = "railway"

# Si no hay variable de entorno (local), poner un string vacío para que no falle al iniciar
# (aunque fallará al intentar conectar si no la pones)
if not DB_PASSWORD:
    DB_PASSWORD = ""

DATABASE_URL = f"postgresql://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}"

engine = create_engine(DATABASE_URL)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
metadata = MetaData()

# Definición de la Tabla Postgres
tabla_stock = Table(
    "stock_items",
    metadata,
    Column("id", Integer, primary_key=True, index=True),
    Column("codigo", String),
    Column("articulo", String),
    Column("stock", String),
    Column("stock_minimo", String),
    Column("stock_optimo", String),
    Column("marca", String),
)

# Crear tabla si no existe
try:
    metadata.create_all(bind=engine)
except Exception as e:
    print(f"Advertencia: No se pudo conectar a Postgres al inicio. {e}")

# ==========================================
# 2. FUNCIONES AUXILIARES (COLORES EXCEL)
# ==========================================

def hex_to_rgb(hex_code):
    if hex_code is None:
        return None
    hex_code = str(hex_code) # Asegurar string
    if len(hex_code) > 6:
        hex_code = hex_code[-6:]
    return tuple(int(hex_code[i:i+2], 16) for i in (0, 2, 4))

def color_distance(c1, c2):
    return sum((a - b) ** 2 for a, b in zip(c1, c2)) ** 0.5

def interpretar_color(color_hex: str):
    if not color_hex:
        return "NO DEFINIDO"

    try:
        color = hex_to_rgb(color_hex)
    except:
        return "NO DEFINIDO"

    verdes = [(0, 255, 0), (0, 176, 80), (144, 238, 144)]
    amarillos = [(255, 255, 0), (255, 230, 0), (255, 255, 153)]
    rojos = [(255, 0, 0), (192, 0, 0), (255, 102, 102)]

    TOLERANCIA = 100

    for verde in verdes:
        if color_distance(color, verde) < TOLERANCIA:
            return "HAY STOCK"
    for amarillo in amarillos:
        if color_distance(color, amarillo) < TOLERANCIA:
            return "PREGUNTAR"
    for rojo in rojos:
        if color_distance(color, rojo) < TOLERANCIA:
            return "NO HAY STOCK"

    return "DESCONOCIDO"


# ==========================================
# 3. ENDPOINTS ANTERIORES
# ==========================================

@app.post("/leer-excel/")
async def leer_excel(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents), data_only=True)

        resultado = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Verificar si la hoja tiene datos
            if ws.max_row < 2: 
                continue
                
            headers = [cell.value for cell in ws[1]]
            hoja_datos = []

            for row in ws.iter_rows(min_row=2):
                fila = {}
                for idx, cell in enumerate(row):
                    if idx >= len(headers): break # Evitar error de índice
                    
                    col_name = headers[idx] if headers[idx] else f"Col_{idx}"
                    valor = cell.value

                    # Detectar color y estado
                    color = None
                    estado = "NO DEFINIDO"
                    if (
                        cell.fill and
                        cell.fill.start_color and
                        hasattr(cell.fill.start_color, 'rgb') and
                        isinstance(cell.fill.start_color.rgb, str)
                    ):
                        color = cell.fill.start_color.rgb
                        estado = interpretar_color(color)

                    # Guardar en la fila
                    fila[str(col_name)] = valor
                    fila[f"{col_name}__color"] = color
                    fila[f"{col_name}__estado"] = estado

                hoja_datos.append(fila)

            resultado[sheet_name] = hoja_datos

        return resultado

    except Exception as e:
        return {
            "error": "Error interno al procesar el archivo",
            "detalle": str(e)
        }

@app.post("/extract")
async def extract(file: UploadFile = File(...)):
    # 1. Guardar archivo temporal
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".sqlite")
    try:
        contents = await file.read()
        tmp.write(contents)
        tmp.close()

        # 2. Conectar a SQLite
        conn = sqlite3.connect(tmp.name)
        
        # 3. Leer tabla productos
        df = pd.read_sql("SELECT * FROM Articulos", conn)
        conn.close()

        # 4. Devolver en JSON
        return df.to_dict(orient="records")
    except Exception as e:
        return {"error": str(e)}
    finally:
        os.unlink(tmp.name) # Borrar archivo temporal

@app.post("/procesar-zip-sqlite/")
async def procesar_zip_sqlite(
    file: UploadFile = File(...), 
    codigos: str = Form(...) 
):
    # Crear carpeta temporal única para esta petición
    temp_dir = tempfile.mkdtemp()
    zip_path = os.path.join(temp_dir, "archivo.zip")
    
    try:
        # 1. Convertir el string de códigos a una lista
        try:
            lista_codigos = json.loads(codigos)
            if not isinstance(lista_codigos, list):
                raise ValueError
        except:
            return {"error": "El campo 'codigos' debe ser un array JSON válido. Ejemplo: ['COD1', 'COD2']"}

        # 2. Guardar el ZIP subido
        with open(zip_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        # 3. Descomprimir el ZIP
        db_path = None
        try:
            with zipfile.ZipFile(zip_path, 'r') as zf:
                zf.extractall(temp_dir)
                
                # Buscar el archivo .sqlite, .db o .sqlite3
                for filename in zf.namelist():
                    if filename.endswith(('.sqlite', '.db', '.sqlite3')):
                        db_path = os.path.join(temp_dir, filename)
                        break
        except zipfile.BadZipFile:
            return {"error": "El archivo no es un ZIP válido."}

        if not db_path:
            return {"error": "No encontré ningún archivo .sqlite o .db dentro del ZIP."}

        # 4. Conectar a SQLite y filtrar
        conn = sqlite3.connect(db_path)
        
        placeholders = ','.join('?' for _ in lista_codigos)
        query = f"SELECT * FROM Articulos WHERE Codigo IN ({placeholders})"
        
        try:
            df = pd.read_sql_query(query, conn, params=lista_codigos)
            resultado = df.to_dict(orient="records")
        except Exception as sql_e:
            return {"error": "Error al consultar la tabla 'Articulos'.", "detalle": str(sql_e)}
        finally:
            conn.close()

        return {
            "mensaje": "Búsqueda exitosa",
            "encontrados": len(resultado),
            "datos": resultado
        }

    except Exception as e:
        return {"error": "Error interno del servidor", "detalle": str(e)}
    
    finally:
        # 5. Limpieza: Borrar todo lo temporal
        shutil.rmtree(temp_dir, ignore_errors=True)


# ==========================================
# 4. NUEVO ENDPOINT (GOOGLE SHEETS -> POSTGRES)
# ==========================================

# Modelo de datos Pydantic
class FilaExcel(BaseModel):
    codigo: Optional[str] = Field(alias="Código", default=None)
    articulo: Optional[str] = Field(alias="Artículo", default=None)
    stock: Optional[str] = Field(alias="Stock", default=None)
    stock_minimo: Optional[str] = Field(alias="Stock Mínimo", default=None)
    stock_optimo: Optional[str] = Field(alias="Stock Optimo", default=None)
    marca: Optional[str] = Field(alias="Marca", default=None)

    class Config:
        populate_by_name = True

def procesar_guardado_postgres(datos: List[FilaExcel]):
    db = SessionLocal()
    try:
        datos_para_db = []
        for fila in datos:
            datos_para_db.append({
                "codigo": fila.codigo,
                "articulo": fila.articulo,
                "stock": fila.stock,
                "stock_minimo": fila.stock_minimo,
                "stock_optimo": fila.stock_optimo,
                "marca": fila.marca
            })
        
        with db.begin():
            db.execute(tabla_stock.delete()) 
            if datos_para_db:
                db.execute(tabla_stock.insert(), datos_para_db) 
        
        print(f"--- ÉXITO: Se sincronizaron {len(datos_para_db)} filas en Postgres ---")
            
    except Exception as e:
        print(f"!!! ERROR CRÍTICO EN POSTGRES: {e}")
    finally:
        db.close()

@app.post("/upload-sheet")
async def recibir_excel_para_postgres(filas: List[FilaExcel], background_tasks: BackgroundTasks):
    print(f"Recibiendo petición /upload-sheet con {len(filas)} filas...")
    
    if not filas:
        raise HTTPException(status_code=400, detail="El Excel parece estar vacío")

    background_tasks.add_task(procesar_guardado_postgres, filas)
    
    return {"message": "Datos recibidos. Procesando en segundo plano."}
